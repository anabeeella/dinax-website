#!/usr/bin/env python3
"""
Script to convert products-list.xlsx to products.json
Install required package: pip install openpyxl
"""

import json
import os
from openpyxl import load_workbook

def convert_xlsx_to_json():
    # Paths
    xlsx_path = "assets/images/products/products-list.xlsx"
    json_path = "data/products.json"
    
    # Check if Excel file exists
    if not os.path.exists(xlsx_path):
        print(f"Error: {xlsx_path} not found!")
        return
    
    # Load workbook
    wb = load_workbook(xlsx_path, data_only=True)
    
    # Get the first sheet
    ws = wb.active
    
    # Read header row (assuming first row contains headers)
    headers = [cell.value for cell in ws[1]]
    print(f"Headers found: {headers}")
    
    # Read all data rows
    products = []
    categories_set = set()
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        # Skip empty rows
        if not any(cell.value for cell in row):
            continue
        
        # Create product object
        product = {}
        
        # Map Excel columns to JSON fields
        for col_idx, header in enumerate(headers):
            if header is None:
                continue
            
            cell_value = row[col_idx].value if col_idx < len(row) else None
            
            if cell_value is None:
                continue
            
            header_lower = str(header).lower().strip()
            
            # Map common field names
            if 'id' in header_lower:
                product['id'] = int(cell_value) if isinstance(cell_value, (int, float)) else cell_value
            elif 'code' in header_lower or 'código' in header_lower:
                product['code'] = str(cell_value)
            elif 'name' in header_lower or 'nombre' in header_lower or 'producto' in header_lower:
                product['name'] = str(cell_value)
            elif 'category' in header_lower or 'categoría' in header_lower:
                category = str(cell_value).lower().strip()
                # Normalize category names to match JSON format
                category_map = {
                    'cámaras y fotografía': 'camera',
                    'camera': 'camera',
                    'productos de belleza': 'beauty',
                    'beauty': 'beauty',
                    'audio y sonido': 'audio',
                    'audio': 'audio',
                    'hogar y cocina': 'home',
                    'hogar': 'home',
                    'home': 'home',
                    'computación y gamer': 'computing',
                    'computing': 'computing',
                    'imagen y tv': 'image',
                    'image': 'image',
                    'electrónica y accesorios': 'electronics',
                    'electrónica': 'electronics',
                    'electronics': 'electronics',
                    'celulares y tablets': 'mobile',
                    'mobile': 'mobile'
                }
                category_normalized = category_map.get(category, category.replace(' ', '-').replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u'))
                product['category'] = category_normalized
                categories_set.add(category_normalized)
            elif 'description' in header_lower or 'descripción' in header_lower:
                product['description'] = str(cell_value)
            elif 'details' in header_lower or 'detalles' in header_lower:
                # If details is a string, try to split by semicolon, comma, or newline
                if isinstance(cell_value, str):
                    separator = ';' if ';' in cell_value else (',' if ',' in cell_value else '\n')
                    details = [d.strip() for d in cell_value.split(separator)]
                    product['details'] = [d for d in details if d]
                else:
                    product['details'] = [str(cell_value)]
            elif 'features' in header_lower or 'características' in header_lower or 'especificaciones' in header_lower:
                # If features is a string, try to split by semicolon, comma, or newline
                if isinstance(cell_value, str):
                    separator = ';' if ';' in cell_value else (',' if ',' in cell_value else '\n')
                    features = [f.strip() for f in cell_value.split(separator)]
                    product['features'] = [f for f in features if f]
                else:
                    product['features'] = [str(cell_value)]
            elif 'image' in header_lower or 'imagen' in header_lower or 'imágenes' in header_lower:
                # Handle images - separated by semicolon (;)
                if isinstance(cell_value, str):
                    # Split by semicolon first, then by comma as fallback
                    separator = ';' if ';' in cell_value else ','
                    images = [img.strip() for img in cell_value.split(separator)]
                    product['images'] = [img for img in images if img]
                else:
                    product['images'] = [str(cell_value)] if cell_value else []
        
        # Add product if it has at least a name
        if product.get('name'):
            # Auto-generate ID if missing
            if 'id' not in product:
                product['id'] = len(products) + 1
            
            # Normalize image paths: add base path and extension if missing
            if 'images' in product and product['images']:
                normalized_images = []
                product_code = product.get('code', '')
                base_path = 'assets/images/products/'
                
                for img in product['images']:
                    if not img:  # Skip empty strings
                        continue
                    
                    # Remove trailing semicolons or spaces
                    img = img.strip().rstrip(';').strip()
                    
                    # Check if image already has a full path
                    if img.startswith('assets/') or img.startswith('/'):
                        # Already has path, just ensure it has extension
                        if '.' not in os.path.basename(img):
                            # No extension found, add .png
                            img = img + '.png'
                        normalized_images.append(img)
                    else:
                        # No path, construct it
                        # If the image name contains the code or matches pattern, use it directly
                        # Otherwise, try to construct from code + image suffix
                        if product_code and (product_code in img or '_' in img or img.isdigit()):
                            # Image name like "DX-IPCAMDIG_main" or "DX-IPCAMDIG0" or "0"
                            if img.startswith(product_code):
                                # Full name like "DX-IPCAMDIG_main"
                                full_path = base_path + img
                            else:
                                # Partial name like "_main" or "0" or "main"
                                if img.startswith('_'):
                                    full_path = base_path + product_code + img
                                elif img.isdigit() or img in ['0', '1', '2', '3', '4', '5']:
                                    full_path = base_path + product_code + img
                                else:
                                    # Just a suffix like "main"
                                    full_path = base_path + product_code + '_' + img
                        else:
                            # Fallback: use the image name as-is
                            full_path = base_path + img
                        
                        # Add extension if missing
                        if '.' not in os.path.basename(full_path):
                            full_path = full_path + '.png'
                        
                        normalized_images.append(full_path)
                
                product['images'] = normalized_images
            
            # If product has no images, add category default image
            if not product.get('images') or len(product.get('images', [])) == 0:
                category = product.get('category', '')
                # Map categories to their default images
                category_image_map = {
                    'audio': 'assets/images/categories/cat-audio.png',
                    'beauty': 'assets/images/categories/cat-beauty.png',
                    'camera': 'assets/images/categories/cat-camera.png',
                    'computing': 'assets/images/categories/cat-computing.png',
                    'electronics': 'assets/images/categories/cat-electronics.png',
                    'home': 'assets/images/categories/cat-home.png',
                    'image': 'assets/images/categories/cat-image.png',
                    'mobile': 'assets/images/categories/cat-cell.png',  # Note: cat-cell instead of cat-mobile
                }
                
                if category in category_image_map:
                    product['images'] = [category_image_map[category]]
                else:
                    # Fallback: use a default image if category not found
                    product['images'] = ['assets/images/categories/cat-electronics.png']
            
            products.append(product)
    
    # Create categories list
    categories = []
    category_names = {
        'camera': 'Cámaras & Fotografía',
        'beauty': 'Productos de Belleza',
        'audio': 'Audio & Sonido',
        'home': 'Hogar & Cocina',
        'computing': 'Computación & Gamer',
        'image': 'Imagen & TV',
        'electronics': 'Electrónica & Accesorios',
        'mobile': 'Celulares & Tablets',
    }
    
    for cat_id in sorted(categories_set):
        categories.append({
            "id": cat_id,
            "name": category_names.get(cat_id, cat_id.capitalize())
        })
    
    # Create final JSON structure
    output = {
        "products": products,
        "categories": categories
    }
    
    # Write to JSON file
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    
    print(f"\n✅ Conversion complete!")
    print(f"📦 {len(products)} products converted")
    print(f"📂 {len(categories)} categories found")
    print(f"💾 Saved to {json_path}")
    
    # Display sample product
    if products:
        print(f"\n📋 Sample product:")
        print(json.dumps(products[0], indent=2, ensure_ascii=False))

if __name__ == "__main__":
    try:
        convert_xlsx_to_json()
    except ImportError:
        print("Error: openpyxl not installed. Install it with: pip install openpyxl")
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

